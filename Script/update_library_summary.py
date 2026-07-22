#!/usr/bin/env python3
"""Synchronize index.json and the venue directory workbook with live 2026 JSONs.

Paper-record schemas are never changed.  Only the top-level index and workbook
summary fields are refreshed from the seven live category directories.
"""

from __future__ import annotations

import datetime as dt
import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent.parent
SOURCE = ROOT / "Source" / "2026"
INDEX_PATH = SOURCE / "index.json"
WORKBOOK_PATH = ROOT / "Top_Venues_Official_Paper_Directory.xlsx"

DISPLAY_NAMES = {
    "SP": "IEEE S&P", "USENIXSec": "USENIX Security", "TDSC": "IEEE TDSC",
    "TIFS": "IEEE TIFS", "TPAMI": "IEEE TPAMI", "AIJ": "Artificial Intelligence",
    "TSE": "IEEE TSE", "TOSEM": "ACM TOSEM", "TOPLAS": "ACM TOPLAS",
}

NOTE_UPDATES = {
    "IEEE S&P": "官方 Accepted Papers 页面；Cycle 1 为 1 篇、Cycle 2 为 253 篇，本地入库共 254 篇。",
    "ACM CCS": "审稿进行中（Cycle B）；会议为 2026 年 11 月，官方论文目录尚未发布。",
    "USENIX Security": "官方 USENIX Technical Sessions 页面；Cycle 1 + Cycle 2，本地入库共 362 篇。",
    "NDSS": "官方 Accepted Papers 页面；Summer Cycle 113 篇 + Fall Cycle 152 篇，本地入库共 265 篇。",
    "IEEE TDSC": "IEEE Xplore Vol. 23（2026，持续出版）；通过 CSV 手工导入，本地 631 篇。自动访问受 HTTP 418 限制。",
    "IEEE TIFS": "IEEE Xplore Vol. 21（2026，持续出版）；通过 CSV 手工导入，本地 443 篇。自动访问受 HTTP 418 限制。",
    "NeurIPS": "NeurIPS 2026 官方页面已建立，但论文目录尚未发布（通常为 12 月会议）。",
    "ICML": "官方 virtual proceedings 动态数据（Orals + Posters，排除 Workshops）；本地入库共 6,796 篇。",
    "ICLR": "官方 virtual proceedings 动态数据（Orals + Posters，排除 Workshops）；本地入库共 5,691 篇。",
    "AAAI": "AAAI-26 OJS Issue Archive；Technical Tracks，本地入库共 2,274 篇。",
    "IJCAI": "官方 Accepted Papers 页面；Main Track 713 篇及其他 Special Tracks，本地入库共 990 篇。",
    "ACL": "官方 ACL 2026 Accepted Papers 页面；Main Conference Long + Short Papers，不含 Findings/Demo/Industry，本地入库共 515 篇。",
    "JMLR": "JMLR Vol. 27（2026，持续出版）；官方 volume 页面，本地入库共 137 篇。",
    "IEEE TPAMI": "IEEE Xplore Vol. 48（2026，持续出版）；通过 CSV 手工导入，本地 631 篇。自动访问受 HTTP 418 限制。",
    "Artificial Intelligence": "Elsevier ScienceDirect 2026 卷期（持续出版）；通过 BibTeX 手工导入，本地 78 篇。自动访问受 HTTP 403 限制。",
    "ICSE": "官方 Research Track 页面；不含其他 tracks，本地入库共 321 篇。",
    "FSE": "官方 Research Papers 页面，本地入库共 211 篇。",
    "ASE": "ASE 2026 官方页面已建立，但论文目录尚未发布。",
    "ISSTA": "官方 Research Papers 页面，本地入库共 176 篇。",
    "IEEE TSE": "IEEE Xplore Vol. 52（2026，持续出版）；通过 CSV 手工导入，本地 109 篇。自动访问受 HTTP 418 限制。",
    "ACM TOSEM": "ACM DL Vol. 35（2026，持续出版）；按月份 BibTeX 手工导入，本地 212 篇。自动访问受 HTTP 403 限制。",
    "PLDI": "官方 PLDI Research Papers 页面，本地入库共 118 篇。",
    "POPL": "官方 POPL Research Papers 页面，本地入库共 92 篇。",
    "OOPSLA": "官方 OOPSLA 2026 页面；Round 1 + Round 2，本地入库共 74 篇。",
    "ACM TOPLAS": "ACM DL Vol. 48（2026，持续出版）；按月份 BibTeX 手工导入，本地 12 篇。自动访问受 HTTP 403 限制。",
    "JACM": "ACM DL Vol. 73（2026，持续出版）；按月份 BibTeX 手工导入，本地 21 篇。自动访问受 HTTP 403 限制。",
    "SOSP": "SOSP 2026 官方页面存在，但论文目录状态为 TBD，尚未发布。",
    "OSDI": "官方 USENIX Technical Sessions 页面，本地入库共 137 篇。",
    "NSDI": "官方 USENIX Technical Sessions 页面，本地入库共 151 篇。",
    "EuroSys": "官方 Papers 页面，本地入库共 138 篇。",
    "ASPLOS": "官方 Program 页面，本地入库共 168 篇。",
    "FAST": "官方 USENIX Technical Sessions 页面，本地入库共 46 篇。",
    "SIGMOD": "官方 SIGMOD Papers 页面；Rounds 1–4、PACMMOD journal-style，本地入库共 311 篇。",
    "VLDB": "PVLDB Vol. 19（2025–2026，持续出版）；本地入库共 81 篇。",
    "ICDE": "官方 ICDE 2026 Research Papers 页面，本地入库共 261 篇。",
    "CAV": "官方 CAV 2026 Accepted Papers 页面；55 Research + 20 Tool + 6 Industrial，本地入库共 81 篇。",
}


def live_files() -> list[Path]:
    return sorted(
        path for path in SOURCE.rglob("*-2026.json")
        if not any(part.startswith("_") for part in path.relative_to(SOURCE).parts)
    )


def collect() -> tuple[dict[str, dict[str, int]], dict[str, int]]:
    per_venue: dict[str, dict[str, int]] = {}
    totals = {"papers": 0, "with_abstract": 0, "without_abstract": 0}
    for path in live_files():
        records = json.loads(path.read_text(encoding="utf-8"))
        if not records:
            continue
        venue = str(records[0]["venue"])
        papers = len(records)
        abstracts = sum(bool(str(record.get("abstract") or "").strip()) for record in records)
        per_venue[venue] = {
            "papers": papers,
            "has_abstract": abstracts,
            "without_abstract": papers - abstracts,
            "has_keywords": sum(bool(record.get("keywords")) for record in records),
            "has_cycle": sum(record.get("cycle") is not None for record in records),
        }
        totals["papers"] += papers
        totals["with_abstract"] += abstracts
        totals["without_abstract"] += papers - abstracts
    return per_venue, totals


def update_index(per_venue: dict[str, dict[str, int]], totals: dict[str, int], today: str) -> None:
    index = json.loads(INDEX_PATH.read_text(encoding="utf-8"))
    for entry in index["venues"]:
        venue = entry.get("file", "").replace("-2026.json", "")
        # Most filename stems equal the record venue code; retain explicit exceptions.
        aliases = {"USENIXSec": "USENIXSec", "AIJ": "AIJ", "SP": "SP"}
        code = aliases.get(venue, venue)
        if code in per_venue:
            stats = per_venue[code]
            entry["paper_count"] = stats["papers"]
            entry["has_abstract"] = stats["has_abstract"]
            entry["has_keywords"] = stats["has_keywords"]
            entry["has_cycle"] = stats["has_cycle"]
    index["last_updated"] = today
    index["total_papers"] = totals["papers"]
    index["abstract_summary"] = {
        "papers_with_abstract": totals["with_abstract"],
        "papers_without_abstract": totals["without_abstract"],
        "coverage_percent": round(totals["with_abstract"] * 100 / totals["papers"], 2),
        "policy": "Official proceedings/pages first; title-verified arXiv fallback permitted for reading-library abstracts.",
    }
    index["live_category_directories"] = ["AI", "DB", "Formal_Methods", "PL", "SE", "Security", "Sys"]
    index["auxiliary_directories_excluded"] = "All underscore-prefixed directories are audit/backup artefacts, not corpus inputs."
    INDEX_PATH.write_text(json.dumps(index, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def set_header(ws, column: int, title: str, source_column: int) -> None:
    cell = ws.cell(1, column, title)
    origin = ws.cell(1, source_column)
    cell._style = copy(origin._style)
    cell.font = copy(origin.font)
    cell.fill = copy(origin.fill)
    cell.border = copy(origin.border)
    cell.alignment = copy(origin.alignment)
    cell.protection = copy(origin.protection)
    cell.comment = Comment("Source: live Source/2026 venue JSON records, refreshed 2026-07-19.", "Codex")


def update_workbook(per_venue: dict[str, dict[str, int]], totals: dict[str, int]) -> None:
    workbook = load_workbook(WORKBOOK_PATH)
    for sheet_name in ("汇总", "2026"):
        ws = workbook[sheet_name]
        headers = {ws.cell(1, column).value: column for column in range(1, ws.max_column + 1)}
        start_column = headers.get("2026本地论文数")
        status_source = 11 if sheet_name == "汇总" else 5
        if start_column is None:
            start_column = ws.max_column + 1
            for offset, header in enumerate(("2026本地论文数", "2026摘要覆盖", "2026摘要状态")):
                set_header(ws, start_column + offset, header, status_source)
        ws.column_dimensions[get_column_letter(start_column)].width = 16
        ws.column_dimensions[get_column_letter(start_column + 1)].width = 19
        ws.column_dimensions[get_column_letter(start_column + 2)].width = 27
        for row in range(2, ws.max_row + 1):
            display = ws.cell(row, 2).value
            note_column = 10 if sheet_name == "汇总" else 6
            if display in NOTE_UPDATES:
                ws.cell(row, note_column, NOTE_UPDATES[display])
            code = next((key for key, value in DISPLAY_NAMES.items() if value == display), str(display or ""))
            # Remaining codes match their workbook display names exactly.
            stats = per_venue.get(code)
            if stats:
                papers = stats["papers"]
                with_abstract = stats["has_abstract"]
                missing = stats["without_abstract"]
                ws.cell(row, start_column, papers)
                ws.cell(row, start_column + 1, f"{with_abstract}/{papers} ({with_abstract * 100 / papers:.1f}%)")
                ws.cell(row, start_column + 2, "摘要已补齐" if not missing else f"仍缺 {missing} 篇摘要")
            elif display in {"ACM CCS", "NeurIPS", "ASE", "SOSP"}:
                ws.cell(row, start_column, 0)
                ws.cell(row, start_column + 1, "—")
                ws.cell(row, start_column + 2, "论文目录尚未发布")
    summary = workbook["汇总"]
    row = next((cell.row for cell in summary["A"] if cell.value == "2026本地库汇总"), summary.max_row + 2)
    summary.cell(row, 1, "2026本地库汇总")
    summary.cell(row, 2, "32 个 venue，21,787 篇论文")
    summary.cell(row, 3, f"摘要：{totals['with_abstract']}/{totals['papers']} ({totals['with_abstract'] * 100 / totals['papers']:.1f}%)；仍缺 {totals['without_abstract']} 篇")
    summary.cell(row, 1).font = copy(summary.cell(1, 1).font)
    summary.cell(row, 2).font = copy(summary.cell(1, 1).font)
    summary.cell(row, 3).font = copy(summary.cell(1, 1).font)
    workbook.save(WORKBOOK_PATH)


def main() -> None:
    today = dt.date.today().isoformat()
    per_venue, totals = collect()
    if len(per_venue) != 32 or totals["papers"] != 21787:
        raise RuntimeError(f"unexpected live corpus: venues={len(per_venue)}, papers={totals['papers']}")
    update_index(per_venue, totals, today)
    update_workbook(per_venue, totals)
    print(json.dumps({"venues": len(per_venue), **totals}, ensure_ascii=False))


if __name__ == "__main__":
    main()
